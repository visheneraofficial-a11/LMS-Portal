"""
ENABLE PROGRAM — Admin Utilities
Shared mixins, actions, and helpers for the enhanced admin console.
"""
import csv
import io
import json
from datetime import datetime, timedelta

from django.contrib import admin, messages
from django.http import HttpResponse, HttpResponseRedirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import format_html
from django.db.models import Count, Q


# ═══════════════════════════════════════════════════════════════════
# EXPORT ACTIONS
# ═══════════════════════════════════════════════════════════════════

def export_as_csv(modeladmin, request, queryset):
    """Export selected records as CSV."""
    meta = modeladmin.model._meta
    field_names = [f.name for f in meta.fields]

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename={meta.verbose_name_plural.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

    writer = csv.writer(response)
    writer.writerow(field_names)

    for obj in queryset:
        row = []
        for field in field_names:
            value = getattr(obj, field)
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            row.append(str(value) if value is not None else '')
        writer.writerow(row)

    modeladmin.message_user(request, f"Exported {queryset.count()} {meta.verbose_name_plural} to CSV.", messages.SUCCESS)
    return response

export_as_csv.short_description = "📥 Export selected as CSV"


def export_as_json(modeladmin, request, queryset):
    """Export selected records as JSON."""
    meta = modeladmin.model._meta
    field_names = [f.name for f in meta.fields]

    data = []
    for obj in queryset:
        row = {}
        for field in field_names:
            value = getattr(obj, field)
            if isinstance(value, datetime):
                value = value.isoformat()
            elif hasattr(value, 'pk'):
                value = str(value.pk)
            else:
                value = str(value) if value is not None else None
            row[field] = value
        data.append(row)

    response = HttpResponse(
        json.dumps(data, indent=2, ensure_ascii=False),
        content_type='application/json'
    )
    response['Content-Disposition'] = f'attachment; filename={meta.verbose_name_plural.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'

    modeladmin.message_user(request, f"Exported {queryset.count()} {meta.verbose_name_plural} to JSON.", messages.SUCCESS)
    return response

export_as_json.short_description = "📥 Export selected as JSON"


# ═══════════════════════════════════════════════════════════════════
# STATUS ACTIONS
# ═══════════════════════════════════════════════════════════════════

def activate_selected(modeladmin, request, queryset):
    """Activate selected records (set status=ACTIVE or is_active=True)."""
    count = 0
    if hasattr(modeladmin.model, 'status'):
        count = queryset.update(status='ACTIVE')
    elif hasattr(modeladmin.model, 'is_active'):
        count = queryset.update(is_active=True)
    modeladmin.message_user(request, f"Activated {count} records.", messages.SUCCESS)

activate_selected.short_description = "✅ Activate selected"


def deactivate_selected(modeladmin, request, queryset):
    """Deactivate selected records."""
    count = 0
    if hasattr(modeladmin.model, 'status'):
        count = queryset.update(status='INACTIVE')
    elif hasattr(modeladmin.model, 'is_active'):
        count = queryset.update(is_active=False)
    modeladmin.message_user(request, f"Deactivated {count} records.", messages.SUCCESS)

deactivate_selected.short_description = "🚫 Deactivate selected"


def suspend_selected(modeladmin, request, queryset):
    """Suspend selected records."""
    if hasattr(modeladmin.model, 'status'):
        count = queryset.update(status='SUSPENDED')
        modeladmin.message_user(request, f"Suspended {count} records.", messages.WARNING)

suspend_selected.short_description = "⏸ Suspend selected"


# ═══════════════════════════════════════════════════════════════════
# DISPLAY HELPERS
# ═══════════════════════════════════════════════════════════════════

def colored_status(status):
    """Returns HTML for a color-coded status badge with light backgrounds."""
    color_map = {
        'ACTIVE': ('#15803d', 'rgba(16, 185, 129, 0.15)'),
        'INACTIVE': ('#dc2626', 'rgba(239, 68, 68, 0.15)'),
        'SUSPENDED': ('#d97706', 'rgba(245, 158, 11, 0.15)'),
        'PENDING_VERIFICATION': ('#2563eb', 'rgba(59, 130, 246, 0.15)'),
        'PENDING': ('#2563eb', 'rgba(59, 130, 246, 0.15)'),
        'DRAFT': ('#64748b', 'rgba(148, 163, 184, 0.15)'),
        'PUBLISHED': ('#15803d', 'rgba(16, 185, 129, 0.15)'),
        'COMPLETED': ('#15803d', 'rgba(16, 185, 129, 0.15)'),
        'CANCELLED': ('#dc2626', 'rgba(239, 68, 68, 0.15)'),
        'IN_PROGRESS': ('#d97706', 'rgba(245, 158, 11, 0.15)'),
        'SCHEDULED': ('#7c3aed', 'rgba(139, 92, 246, 0.15)'),
        'LIVE': ('#dc2626', 'rgba(239, 68, 68, 0.15)'),
        'OPEN': ('#15803d', 'rgba(16, 185, 129, 0.15)'),
        'CLOSED': ('#64748b', 'rgba(148, 163, 184, 0.15)'),
        'RESOLVED': ('#15803d', 'rgba(16, 185, 129, 0.15)'),
        'PAID': ('#15803d', 'rgba(16, 185, 129, 0.15)'),
        'UNPAID': ('#dc2626', 'rgba(239, 68, 68, 0.15)'),
        'PARTIAL': ('#d97706', 'rgba(245, 158, 11, 0.15)'),
        True: ('#15803d', 'rgba(16, 185, 129, 0.15)'),
        False: ('#dc2626', 'rgba(239, 68, 68, 0.15)'),
    }
    fg, bg = color_map.get(status, ('#64748b', 'rgba(148, 163, 184, 0.15)'))
    label = str(status).replace('_', ' ').title() if isinstance(status, str) else ('Active' if status else 'Inactive')
    return format_html(
        '<span style="display:inline-flex;align-items:center;padding:4px 12px;'
        'border-radius:6px;font-size:0.72rem;font-weight:600;'
        'color:{};background:{};letter-spacing:0.03em;'
        'border:1px solid {};">{}</span>',
        fg, bg, fg, label
    )


# ═══════════════════════════════════════════════════════════════════
# ENHANCED BASE ADMIN
# ═══════════════════════════════════════════════════════════════════

class EnhancedModelAdmin(admin.ModelAdmin):
    """Base ModelAdmin with export actions, colored statuses, and enhanced UI."""

    actions = [export_as_csv, export_as_json, activate_selected, deactivate_selected]
    list_per_page = 25
    show_full_result_count = True
    save_on_top = True

    class Media:
        css = {'all': ('css/admin_theme.css',)}

    def changelist_view(self, request, extra_context=None):
        """Override to support dynamic list_per_page from query parameter."""
        per_page = request.GET.get('list_per_page')
        if per_page:
            try:
                per_page_val = int(per_page)
                if per_page_val in (10, 25, 50, 100):
                    self.list_per_page = per_page_val
            except (ValueError, TypeError):
                pass
            # Remove list_per_page from GET so Django doesn't treat it as a filter
            modified_get = request.GET.copy()
            modified_get.pop('list_per_page', None)
            request.GET = modified_get
        return super().changelist_view(request, extra_context=extra_context)

    def status_badge(self, obj):
        status = getattr(obj, 'status', None)
        if status is None:
            status = getattr(obj, 'is_active', None)
        if status is not None:
            return colored_status(status)
        return '-'
    status_badge.short_description = 'Status'
    status_badge.admin_order_field = 'status'

    def created_display(self, obj):
        dt = getattr(obj, 'created_at', None)
        if dt:
            return format_html(
                '<span style="color:#94a3b8;font-size:0.82rem;" title="{}">{}</span>',
                dt.strftime('%Y-%m-%d %H:%M:%S'),
                dt.strftime('%d %b %Y')
            )
        return '-'
    created_display.short_description = 'Created'
    created_display.admin_order_field = 'created_at'

    def updated_display(self, obj):
        dt = getattr(obj, 'updated_at', None)
        if dt:
            return format_html(
                '<span style="color:#94a3b8;font-size:0.82rem;" title="{}">{}</span>',
                dt.strftime('%Y-%m-%d %H:%M:%S'),
                dt.strftime('%d %b %Y')
            )
        return '-'
    updated_display.short_description = 'Updated'
    updated_display.admin_order_field = 'updated_at'


# ═══════════════════════════════════════════════════════════════════
# IMPORT MIXIN
# ═══════════════════════════════════════════════════════════════════

class ImportExportMixin:
    """Adds import/export URL endpoints to a ModelAdmin."""
    change_list_template = 'admin/enhanced_change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-csv/', self.import_csv_view, name=f'{self.model._meta.app_label}_{self.model._meta.model_name}_import_csv'),
            path('export-all-csv/', self.export_all_csv_view, name=f'{self.model._meta.app_label}_{self.model._meta.model_name}_export_all_csv'),
            path('export-filtered-excel/', self.export_filtered_excel_view, name=f'{self.model._meta.app_label}_{self.model._meta.model_name}_export_filtered_excel'),
        ]
        return custom_urls + urls

    def import_csv_view(self, request):
        if request.method == 'POST' and request.FILES.get('csv_file'):
            csv_file = request.FILES['csv_file']
            try:
                decoded = csv_file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                created_count = 0
                error_count = 0
                for row in reader:
                    try:
                        # Clean row: strip whitespace from keys and values
                        clean_row = {k.strip(): v.strip() if isinstance(v, str) else v for k, v in row.items() if k}
                        # Remove id/pk/uuid fields for creation
                        for remove_key in ['id', 'pk', 'uuid', 'created_at', 'updated_at']:
                            clean_row.pop(remove_key, None)

                        # Get valid field names
                        valid_fields = {f.name for f in self.model._meta.fields}
                        filtered_row = {k: v for k, v in clean_row.items() if k in valid_fields}

                        self.model.objects.create(**filtered_row)
                        created_count += 1
                    except Exception as e:
                        error_count += 1

                self.message_user(
                    request,
                    f"Import complete: {created_count} created, {error_count} errors.",
                    messages.SUCCESS if error_count == 0 else messages.WARNING
                )
            except Exception as e:
                self.message_user(request, f"Import failed: {str(e)}", messages.ERROR)

            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '../'))

        context = {
            **self.admin_site.each_context(request),
            'title': f'Import {self.model._meta.verbose_name_plural.title()}',
            'model_name': self.model._meta.verbose_name_plural.title(),
            'opts': self.model._meta,
        }
        return TemplateResponse(request, 'admin/import_csv.html', context)

    def export_all_csv_view(self, request):
        queryset = self.model.objects.all()
        return export_as_csv(self, request, queryset)

    def export_filtered_excel_view(self, request):
        """Export currently filtered/displayed records as Excel (.xlsx)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self.message_user(request, "openpyxl is required for Excel export. Please install it.", messages.ERROR)
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', '../'))

        from django.contrib.admin.views.main import ChangeList

        # Get the filtered queryset via ChangeList (same as Django's changelist_view)
        list_display = self.get_list_display(request)
        list_display_links = self.get_list_display_links(request, list_display)
        list_filter = self.get_list_filter(request)
        search_fields = self.get_search_fields(request)
        list_select_related = self.get_list_select_related(request)

        try:
            cl = ChangeList(
                request,
                self.model,
                list_display,
                list_display_links,
                list_filter,
                self.date_hierarchy,
                search_fields,
                list_select_related,
                self.list_per_page,
                self.list_max_show_all,
                self.list_editable,
                self,
                self.sortable_by,
                self.search_help_text,
            )
            # get_queryset applies search, filters, and ordering — NOT pagination
            queryset = cl.get_queryset(request)
        except Exception:
            # Fallback: manually apply filters
            queryset = self.get_queryset(request)
            search_term = request.GET.get('q', '').strip()
            if search_term and self.search_fields:
                search_query = Q()
                for field in self.search_fields:
                    search_query |= Q(**{f"{field}__icontains": search_term})
                queryset = queryset.filter(search_query)
            for key, value in request.GET.items():
                if key in ('q', 'p', 'o', 'list_per_page', '_changelist_filters', 'e', 'all'):
                    continue
                if value:
                    field_name = key.replace('__exact', '').replace('__in', '').replace('__isnull', '')
                    try:
                        self.model._meta.get_field(field_name)
                        queryset = queryset.filter(**{key: value})
                    except Exception:
                        pass

        meta = self.model._meta

        # Get field names from list_display
        field_names = []
        method_names = []
        for name in list_display:
            if hasattr(self.model, name):
                field_names.append(name)
            elif hasattr(self, name):
                method_names.append(name)
            else:
                try:
                    self.model._meta.get_field(name)
                    field_names.append(name)
                except Exception:
                    pass

        if not field_names:
            field_names = [f.name for f in meta.fields]

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = meta.verbose_name_plural.title()[:31]

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="844FC1", end_color="844FC1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        headers = []
        for name in field_names:
            try:
                field = meta.get_field(name)
                headers.append(field.verbose_name.title() if hasattr(field, 'verbose_name') else name)
            except Exception:
                headers.append(name.replace('_', ' ').title())
        for name in method_names:
            method = getattr(self, name, None)
            if hasattr(method, 'short_description'):
                headers.append(method.short_description)
            else:
                headers.append(name.replace('_', ' ').title())

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row_num, obj in enumerate(queryset, 2):
            col_num = 1
            for name in field_names:
                value = getattr(obj, name, '')
                if hasattr(value, '__call__'):
                    value = value()
                if value is None:
                    value = ''
                elif hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'hour') else value.strftime('%Y-%m-%d')
                else:
                    value = str(value) if value else ''
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                col_num += 1

            for name in method_names:
                method = getattr(self, name, None)
                if method:
                    try:
                        from django.utils.html import strip_tags
                        value = strip_tags(str(method(obj)))
                    except Exception:
                        value = ''
                else:
                    value = ''
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                col_num += 1

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'{meta.verbose_name_plural.replace(" ", "_")}_filtered_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)

        self.message_user(request, f"Exported {queryset.count()} filtered {meta.verbose_name_plural} to Excel.", messages.SUCCESS)
        return response
